"""
Sonderaufgaben-Widget (portiert aus Nesk2)
Tägliches Aufgaben-Formular mit Dienstplan-Integration und Excel-Export.

Vorlage: Daten/Sonderaufgaben/Sonderaufgaben.xlsx
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime
from pathlib import Path
from shutil import copy2

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QComboBox, QLineEdit, QTextEdit,
    QFrame, QScrollArea, QSizePolicy, QMessageBox, QFileDialog,
    QGroupBox, QTreeView, QSplitter, QFileSystemModel, QMenu,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont

from config import (
    BASE_DIR,
    FIORI_BLUE, FIORI_TEXT, FIORI_WHITE, FIORI_BORDER,
    FIORI_SUCCESS, FIORI_WARNING,
)
from functions.settings_functions import get_setting

# ── Pfad zur Excel-Vorlage ─────────────────────────────────────────────────
TEMPLATE_PATH = Path(BASE_DIR) / "Daten" / "Sonderaufgaben" / "Sonderaufgaben.xlsx"

# Aufgaben-Zeilen im Excel-Template (Zeile → Aufgabenname)
# Tagdienst → Spalte C (3), Nachtdienst → Spalte E (5)
_AUFGABEN_MAPPING = {
    "Sauberkeit Station":      3,
    "BTW Check + Sauberkeit":  4,
    "E-mobby Check":           5,
    "Bulmor 1 - 7312":         6,
    "Bulmor 2 - 7892":         7,
    "Bulmor 3 - 8092":         8,
    "Bulmor 4 - 8794":         9,
    "Bulmor 5 - 9982":         10,
}

# Service-Point-Mapping: key → (Excel-Zeile, Excel-Spalte)
_SERVICE_MAPPING = {
    "C72_06_12": (16, 3),
    "C72_12_18": (16, 4),
    "C72_18_00": (16, 5),
    "C72_00_06": (16, 6),
}

_BULMOR_AUFGABEN = {
    "Bulmor 1 - 7312",
    "Bulmor 2 - 7892",
    "Bulmor 3 - 8092",
    "Bulmor 4 - 8794",
    "Bulmor 5 - 9982",
}

# ── Hilfsfunktionen ─────────────────────────────────────────────────────────

def _btn_style(color: str, hover: str) -> str:
    return f"""
        QPushButton {{
            background-color: {color};
            color: white;
            border: none;
            border-radius: 4px;
            padding: 6px 14px;
            font-weight: bold;
            font-size: 12px;
        }}
        QPushButton:hover {{
            background-color: {hover};
        }}
        QPushButton:pressed {{
            background-color: {hover};
        }}
    """


def _section_label(text: str) -> QLabel:
    lbl = QLabel(text)
    lbl.setFont(QFont("Arial", 12, QFont.Weight.Bold))
    lbl.setStyleSheet(f"color: {FIORI_TEXT}; padding: 4px 0;")
    return lbl


def _combo_style() -> str:
    return """
        QComboBox {
            border: 1px solid #c8d2dc;
            border-radius: 3px;
            padding: 4px 8px;
            font-size: 11px;
            background: white;
        }
        QComboBox::drop-down {
            border: none;
            width: 20px;
        }
        QComboBox:focus {
            border: 1px solid #0a6ed1;
        }
    """


def _line_style() -> str:
    return """
        QLineEdit {
            border: 1px solid #c8d2dc;
            border-radius: 3px;
            padding: 4px 8px;
            font-size: 11px;
            background: white;
        }
        QLineEdit:focus {
            border: 1px solid #0a6ed1;
        }
    """


class SonderaufgabenWidget(QWidget):
    """Sonderaufgaben-Formular – portiert aus Nesk2, angepasst für PySide6."""

    def __init__(self, parent=None):
        super().__init__(parent)

        # Status-Variablen
        self._entries: dict = {}          # key → {'combo': QComboBox, 'line': QLineEdit, 'nur_bulmor': bool}
        self._tag_mitarbeiter:   list[str] = []
        self._nacht_mitarbeiter: list[str] = []
        self._tag_bulmor:        list[str] = []
        self._nacht_bulmor:      list[str] = []
        self._tag_emobby:        list[str] = []
        self._nacht_emobby:      list[str] = []
        self._dienstplan_geladen: bool = False
        self._dienstplan_pfad:   str  = ""
        self._fs_model: QFileSystemModel | None = None

        self._build_ui()

    # ── UI aufbauen ──────────────────────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Header-Bereich ──────────────────────────────────────────────────
        header_frame = QFrame()
        header_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {FIORI_WHITE};
                border-bottom: 1px solid {FIORI_BORDER};
            }}
        """)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(24, 14, 24, 14)

        title = QLabel("📝 Sonderaufgaben")
        title.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        title.setStyleSheet(f"color: {FIORI_TEXT};")
        header_layout.addWidget(title)
        header_layout.addStretch()

        # Buttons
        btn_new = QPushButton("✚ Neue Aufgaben")
        btn_new.setStyleSheet(_btn_style("#495057", "#343a40"))
        btn_new.setMinimumWidth(140)
        btn_new.clicked.connect(self._new_form)
        header_layout.addWidget(btn_new)

        btn_load = QPushButton("📂 Dienstplan laden")
        btn_load.setStyleSheet(_btn_style(FIORI_BLUE, "#0854a7"))
        btn_load.setMinimumWidth(160)
        btn_load.clicked.connect(self._load_dienstplan)
        header_layout.addWidget(btn_load)

        self._btn_open_dienstplan = QPushButton("📋 Dienstplan öffnen")
        self._btn_open_dienstplan.setStyleSheet(_btn_style("#17a2b8", "#117a8b"))
        self._btn_open_dienstplan.setMinimumWidth(160)
        self._btn_open_dienstplan.setEnabled(False)
        self._btn_open_dienstplan.clicked.connect(self._open_dienstplan_excel)
        header_layout.addWidget(self._btn_open_dienstplan)

        btn_save = QPushButton("💾 Speichern")
        btn_save.setStyleSheet(_btn_style(FIORI_SUCCESS, "#0a5c2f"))
        btn_save.setMinimumWidth(120)
        btn_save.clicked.connect(self._save)
        header_layout.addWidget(btn_save)

        btn_open = QPushButton("📄 Excel öffnen")
        btn_open.setStyleSheet(_btn_style("#17a2b8", "#117a8b"))
        btn_open.setMinimumWidth(130)
        btn_open.clicked.connect(self._open_excel)
        header_layout.addWidget(btn_open)

        btn_print = QPushButton("🖨️ Drucken")
        btn_print.setStyleSheet(_btn_style("#6c757d", "#545b62"))
        btn_print.setMinimumWidth(110)
        btn_print.clicked.connect(self._print)
        header_layout.addWidget(btn_print)

        root.addWidget(header_frame)

        # ── Splitter: Dateibaum links | Formular rechts ─────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(4)
        splitter.setStyleSheet("QSplitter::handle { background: #d0d8e4; }")

        # — Linke Seite: Dateibaum ——————————————————————————————————————————
        tree_panel = QWidget()
        tree_panel.setMinimumWidth(170)
        tree_panel.setMaximumWidth(320)
        tree_panel.setStyleSheet("background: #f8f9fa;")
        tree_vbox = QVBoxLayout(tree_panel)
        tree_vbox.setContentsMargins(8, 12, 8, 8)
        tree_vbox.setSpacing(6)

        tree_hdr = QHBoxLayout()
        tree_lbl = QLabel("📁 Gespeicherte Aufgaben")
        tree_lbl.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        tree_lbl.setStyleSheet(f"color: {FIORI_TEXT};")
        tree_hdr.addWidget(tree_lbl)
        tree_hdr.addStretch()
        btn_reload_tree = QPushButton("🔄")
        btn_reload_tree.setFixedSize(26, 26)
        btn_reload_tree.setToolTip("Ordner aktualisieren")
        btn_reload_tree.setStyleSheet("""
            QPushButton { background: transparent; border: 1px solid #ccc;
                          border-radius: 4px; font-size: 12px; }
            QPushButton:hover { background: #e8edf2; }
        """)
        btn_reload_tree.clicked.connect(self.reload_tree)
        tree_hdr.addWidget(btn_reload_tree)
        tree_vbox.addLayout(tree_hdr)

        self._tree = QTreeView()
        self._tree.setStyleSheet("""
            QTreeView {
                background-color: white;
                border: 1px solid #dce8f5;
                border-radius: 6px;
                font-size: 12px;
            }
            QTreeView::item { padding: 4px 2px; }
            QTreeView::item:selected { background-color: #dce8f5; color: #0a5ba4; }
            QTreeView::item:hover    { background-color: #f0f4f8; }
        """)
        self._tree.setAnimated(True)
        self._tree.setSortingEnabled(True)
        self._tree.activated.connect(self._on_tree_activated)
        self._tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._tree.customContextMenuRequested.connect(self._tree_context_menu)
        tree_vbox.addWidget(self._tree, 1)

        self._tree_info_lbl = QLabel("")
        self._tree_info_lbl.setWordWrap(True)
        self._tree_info_lbl.setStyleSheet("color: #aaa; font-size: 9px; padding: 2px;")
        tree_vbox.addWidget(self._tree_info_lbl)

        splitter.addWidget(tree_panel)

        # — Rechte Seite: Scrollbares Formular ——————————————————————————————
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: #f5f6f7; }")

        self._content = QWidget()
        self._content.setStyleSheet("background: #f5f6f7;")
        self._content_layout = QVBoxLayout(self._content)
        self._content_layout.setContentsMargins(24, 20, 24, 24)
        self._content_layout.setSpacing(16)

        scroll.setWidget(self._content)
        splitter.addWidget(scroll)
        splitter.setSizes([220, 900])

        root.addWidget(splitter, 1)

        # Formular + Baum initial aufbauen
        self._build_form()
        self._setup_tree()

    def _build_form(self):
        """Formular aufbauen / neu aufbauen."""
        # Vorhandene Widgets entfernen
        while self._content_layout.count():
            item = self._content_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        self._entries.clear()
        datum_heute = datetime.now().strftime("%d.%m.%Y")

        # Datum-Info
        datum_lbl = QLabel(f"DRK KV Köln e.V.  –  Sonderaufgaben  –  Sanitätsstation CGN        Datum: {datum_heute}")
        datum_lbl.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        datum_lbl.setStyleSheet(f"color: {FIORI_TEXT}; padding: 4px 0 8px 0;")
        self._content_layout.addWidget(datum_lbl)

        # ── Sektion 1: Stationsaufgaben ─────────────────────────────────────
        sec1 = self._make_section("🏥 Stationsaufgaben")
        grid1 = self._make_grid_in_section(sec1)
        self._add_header_row(grid1, 0)
        for i, name in enumerate(["Sauberkeit Station", "BTW Check + Sauberkeit", "E-mobby Check"], start=1):
            self._add_aufgabe_row(grid1, name, i, nur_bulmor=False)
        self._content_layout.addWidget(sec1)

        # ── Sektion 2: Bulmor-Fahrten ────────────────────────────────────────
        sec2 = self._make_section("🚗 Bulmor-Fahrten (nur Bulmor-Fahrer)")
        grid2 = self._make_grid_in_section(sec2)
        self._add_header_row(grid2, 0, show_status=True)
        for i, name in enumerate(
            ["Bulmor 1 - 7312", "Bulmor 2 - 7892", "Bulmor 3 - 8092",
             "Bulmor 4 - 8794", "Bulmor 5 - 9982"], start=1
        ):
            self._add_aufgabe_row(grid2, name, i, nur_bulmor=True)
        self._content_layout.addWidget(sec2)

        # ── Sektion 3: Bemerkung ─────────────────────────────────────────────
        sec3 = self._make_section("💬 Bemerkung")
        sec3_layout = sec3.layout()
        self._bemerkung = QTextEdit()
        self._bemerkung.setFixedHeight(90)
        self._bemerkung.setStyleSheet("""
            QTextEdit {
                border: 1px solid #c8d2dc;
                border-radius: 3px;
                padding: 6px;
                font-size: 12px;
                background: white;
            }
            QTextEdit:focus { border: 1px solid #0a6ed1; }
        """)
        sec3_layout.addWidget(self._bemerkung)
        self._content_layout.addWidget(sec3)

        # ── Sektion 4: Service Point C72 ─────────────────────────────────────
        sec4 = self._make_section("🕐 Service Point Zeiten – MP C72")
        grid4 = self._make_grid_in_section(sec4)

        slots = [("06–12 Uhr", "C72_06_12", True),
                 ("12–18 Uhr", "C72_12_18", True),
                 ("18–00 Uhr", "C72_18_00", False),
                 ("00–06 Uhr", "C72_00_06", False)]

        # Header
        for col, (lbl_text, _, _is_tag) in enumerate(slots):
            h = QLabel(lbl_text)
            h.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            h.setAlignment(Qt.AlignmentFlag.AlignCenter)
            h.setStyleSheet(f"color: {FIORI_TEXT};")
            grid4.addWidget(h, 0, col + 1)

        lbl_c72 = QLabel("MP C72")
        lbl_c72.setFont(QFont("Arial", 11))
        lbl_c72.setStyleSheet(f"color: {FIORI_TEXT};")
        grid4.addWidget(lbl_c72, 1, 0)

        for col, (_, key, is_tag) in enumerate(slots):
            cell = QWidget()
            cell_v = QVBoxLayout(cell)
            cell_v.setContentsMargins(2, 2, 2, 2)
            cell_v.setSpacing(3)

            combo = QComboBox()
            combo.setFixedHeight(28)
            combo.setStyleSheet(_combo_style())
            mitarbeiter = self._tag_mitarbeiter if is_tag else self._nacht_mitarbeiter
            combo.addItems(["— bitte wählen —"] + mitarbeiter if mitarbeiter else ["— Dienstplan laden —"])

            line = QLineEdit()
            line.setFixedHeight(28)
            line.setStyleSheet(_line_style())
            line.setPlaceholderText("Name")

            combo.currentTextChanged.connect(
                lambda txt, k=key: self._combo_to_line(k, txt)
            )

            cell_v.addWidget(combo)
            cell_v.addWidget(line)
            grid4.addWidget(cell, 1, col + 1)

            self._entries[key] = {"combo": combo, "line": line, "nur_bulmor": False}

        self._content_layout.addWidget(sec4)
        self._content_layout.addStretch()

    # ── Formular-Hilfsmethoden ───────────────────────────────────────────────

    def _make_section(self, title: str) -> QGroupBox:
        box = QGroupBox(title)
        box.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        box.setStyleSheet(f"""
            QGroupBox {{
                background: white;
                border: 1px solid {FIORI_BORDER};
                border-radius: 6px;
                margin-top: 8px;
                padding: 10px;
                color: {FIORI_TEXT};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 4px;
            }}
        """)
        layout = QVBoxLayout(box)
        layout.setContentsMargins(10, 16, 10, 10)
        return box

    def _make_grid_in_section(self, section: QGroupBox) -> QGridLayout:
        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(6)
        grid.setColumnStretch(0, 2)
        grid.setColumnStretch(1, 2)
        grid.setColumnStretch(2, 2)
        grid.setColumnStretch(3, 2)
        grid.setColumnStretch(4, 2)
        section.layout().addLayout(grid)
        return grid

    def _add_header_row(self, grid: QGridLayout, row: int, show_status: bool = False):
        headers = ["Aufgabe", "Tagdienst (Auswahl)", "Tagdienst (Name)", "Nachtdienst (Auswahl)", "Nachtdienst (Name)"]
        if show_status:
            headers.append("Fahrzeugstatus")
        for col, text in enumerate(headers):
            lbl = QLabel(text)
            lbl.setFont(QFont("Arial", 10, QFont.Weight.Bold))
            lbl.setStyleSheet(f"color: #666; border-bottom: 1px solid {FIORI_BORDER}; padding-bottom: 4px;")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter if col > 0 else Qt.AlignmentFlag.AlignLeft)
            grid.addWidget(lbl, row, col)

    def _add_aufgabe_row(self, grid: QGridLayout, name: str, row: int, nur_bulmor: bool):
        """Eine Aufgaben-Zeile mit je Dropdown + Freitext für Tag und Nacht."""
        is_emobby = (name == "E-mobby Check")

        # Aufgaben-Label
        lbl = QLabel(name)
        lbl.setFont(QFont("Arial", 11))
        lbl.setStyleSheet(f"color: {FIORI_TEXT};")
        grid.addWidget(lbl, row, 0)

        for schicht in ("tag", "nacht"):
            key = f"{name}_{schicht}"
            is_tag = (schicht == "tag")

            if nur_bulmor:
                mitarbeiter = self._tag_bulmor if is_tag else self._nacht_bulmor
            elif is_emobby:
                mitarbeiter = self._tag_emobby if is_tag else self._nacht_emobby
            else:
                mitarbeiter = self._tag_mitarbeiter if is_tag else self._nacht_mitarbeiter

            combo = QComboBox()
            combo.setFixedHeight(30)
            combo.setStyleSheet(_combo_style())
            if mitarbeiter:
                items = ["\u2014 bitte wählen \u2014"] + mitarbeiter
                if nur_bulmor:
                    items.append("a.D.")
                combo.addItems(items)
            elif is_emobby and self._dienstplan_geladen:
                combo.addItems(["\u26a0 Kein E-Mobby-Fahrer auf dieser Schicht \u2013 bitte prüfen!"])
                combo.setStyleSheet(_combo_style() + "QComboBox { color: #cc6600; font-weight: bold; }")
            else:
                base_items = ["\u2014 Dienstplan laden \u2014"]
                if nur_bulmor:
                    base_items.append("a.D.")
                combo.addItems(base_items)

            line = QLineEdit()
            line.setFixedHeight(30)
            line.setStyleSheet(_line_style())
            line.setPlaceholderText("Tag" if is_tag else "Nacht")

            combo.currentTextChanged.connect(
                lambda txt, k=key: self._combo_to_line(k, txt)
            )

            col_combo = 1 if is_tag else 3
            col_line  = 2 if is_tag else 4
            grid.addWidget(combo, row, col_combo)
            grid.addWidget(line,  row, col_line)

            self._entries[key] = {"combo": combo, "line": line, "nur_bulmor": nur_bulmor}

        # Fahrzeugstatus-Badge für Bulmor-Zeilen
        if nur_bulmor:
            kennzeichen = name.split("-")[-1].strip() if "-" in name else ""
            status_lbl = QLabel(self._bulmor_status_text(kennzeichen))
            status_lbl.setFont(QFont("Arial", 11))
            status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            status_lbl.setStyleSheet(self._bulmor_status_style(kennzeichen))
            status_lbl.setFixedHeight(30)
            grid.addWidget(status_lbl, row, 5)

    def _bulmor_status_text(self, kennzeichen: str) -> str:
        """Gibt den angezeigten Status-Text für ein Bulmor-Fahrzeug zurück."""
        try:
            from functions.fahrzeug_functions import lade_alle_fahrzeuge
            fahrzeuge = lade_alle_fahrzeuge()
            for fz in fahrzeuge:
                if kennzeichen and kennzeichen in str(fz.get("kennzeichen", "")):
                    st = fz.get("aktueller_status") or ""
                    mapping = {
                        "fahrbereit":   "🟢 fahrbereit",
                        "defekt":       "🔴 defekt",
                        "werkstatt":    "🟡 Werkstatt",
                        "ausser_dienst": "⚫ a.D.",
                        "sonstiges":    "🔵 sonstiges",
                    }
                    return mapping.get(st, f"❓ {st}" if st else "❓ unbekannt")
        except Exception:
            pass
        return "❓ unbekannt"

    def _bulmor_status_style(self, kennzeichen: str) -> str:
        """Gibt den CSS-Style passend zum Fahrzeugstatus zurück."""
        try:
            from functions.fahrzeug_functions import lade_alle_fahrzeuge
            fahrzeuge = lade_alle_fahrzeuge()
            for fz in fahrzeuge:
                if kennzeichen and kennzeichen in str(fz.get("kennzeichen", "")):
                    st = fz.get("aktueller_status") or ""
                    color_map = {
                        "fahrbereit":    ("#e8f5e9", "#2e7d32"),
                        "defekt":        ("#ffebee", "#c62828"),
                        "werkstatt":     ("#fff8e1", "#f57f17"),
                        "ausser_dienst": ("#eeeeee", "#424242"),
                        "sonstiges":     ("#e3f2fd", "#1565c0"),
                    }
                    bg, fg = color_map.get(st, ("#f5f5f5", "#888"))
                    return (f"background:{bg}; color:{fg}; border-radius:4px;"
                            f" padding:2px 6px; font-weight:bold;")
        except Exception:
            pass
        return "color:#888; font-style:italic;"
        """Dropdown-Auswahl in Textfeld übertragen. Bulmor: mehrere mit / anhängen."""
        if not choice or choice in ("— bitte wählen —", "— Dienstplan laden —") \
                or choice.startswith("⚠"):
            return
        entry = self._entries.get(key)
        if not entry:
            return
        line: QLineEdit = entry["line"]
        ist_bulmor = entry.get("nur_bulmor", False)

        current = line.text().strip()
        if ist_bulmor and current and current not in ("Tag", "Nacht"):
            line.setText(f"{current} / {choice}")
        else:
            line.setText(choice)

    # ── Dienstplan laden ────────────────────────────────────────────────────

    def _load_dienstplan(self):
        """Dienstplan-Datei per Dialog auswählen und Mitarbeiter laden."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            QMessageBox.critical(self, "Fehler", "openpyxl nicht installiert!\npip install openpyxl")
            return

        initial_dir = get_setting("dienstplan_ordner", "")
        filepath, _ = QFileDialog.getOpenFileName(
            self,
            "Dienstplan auswählen",
            initial_dir,
            "Excel-Dateien (*.xlsx *.xls)"
        )
        if not filepath:
            return
        self._dienstplan_pfad = filepath
        self._parse_dienstplan(filepath)

    def _open_dienstplan_excel(self):
        """Geladene Dienstplan-Datei in Excel öffnen."""
        if not self._dienstplan_pfad or not os.path.isfile(self._dienstplan_pfad):
            QMessageBox.warning(self, "Nicht verfügbar", "Kein Dienstplan geladen oder Datei nicht gefunden.")
            return
        try:
            os.startfile(self._dienstplan_pfad)
        except Exception as exc:
            QMessageBox.critical(self, "Fehler", f"Datei konnte nicht geöffnet werden:\n{exc}")

    def _parse_dienstplan(self, filepath: str):
        """Mitarbeiter aus Dienstplan-Excel extrahieren und Dropdowns befüllen."""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "Fehler", "openpyxl nicht installiert!\npip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(filepath, data_only=False)
            ws = wb.active

            # Header-Zeile mit Name + Dienst finden
            name_col = dienst_col = header_row = None
            for row_idx in range(1, min(20, ws.max_row + 1)):
                for col_idx in range(1, 15):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        val = cell.value.strip().upper()
                        if val == "NAME":
                            name_col = col_idx
                            header_row = row_idx
                        elif val == "DIENST":
                            dienst_col = col_idx
                if name_col and dienst_col:
                    break

            if not name_col or not dienst_col:
                QMessageBox.critical(
                    self, "Fehler",
                    "Konnte Header-Zeile nicht finden.\n"
                    "Benötigt werden Spalten 'Name' und 'Dienst'."
                )
                return

            tag_mitarbeiter:   list[str] = []
            nacht_mitarbeiter: list[str] = []
            tag_bulmor:        list[str] = []
            nacht_bulmor:      list[str] = []

            skip_words = ["name", "mitarbeiter", "datum", "uhrzeit", "von", "bis",
                          "dienst", "schicht", "pause", "station", "terminal", "funktion"]

            for row_idx in range(header_row + 1, min(header_row + 150, ws.max_row + 1)):
                name_cell   = ws.cell(row=row_idx, column=name_col)
                dienst_cell = ws.cell(row=row_idx, column=dienst_col)

                if not (name_cell.value and isinstance(name_cell.value, str)):
                    continue

                name = str(name_cell.value).strip()
                if len(name) < 3 or not any(c.isalpha() for c in name):
                    continue
                if any(w in name.lower() for w in skip_words):
                    continue

                # Bulmor: gelbe Zellfarbe erkennen
                is_bulmor = False
                try:
                    fill = dienst_cell.fill
                    if fill and fill.start_color:
                        rgb = fill.start_color.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            if len(rgb) == 8:
                                rgb = rgb[2:]
                            r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
                            if r > 200 and g > 200 and b < 100:
                                is_bulmor = True
                except Exception:
                    pass

                # Dienstart ermitteln
                dienstart = None
                if dienst_cell.value and isinstance(dienst_cell.value, str):
                    val = str(dienst_cell.value).strip().upper()
                    if val in ("T", "T10", "DT", "DT3", "DT10"):
                        dienstart = "tag"
                    elif val in ("N", "NF", "N10", "DN", "DN3", "DN10"):
                        dienstart = "nacht"

                if dienstart is None:
                    continue

                # Nachname extrahieren
                if "," in name:
                    nachname = name.split(",")[0].strip()
                else:
                    nachname = name.split()[0].strip() if " " in name else name

                if dienstart == "tag" and nachname not in tag_mitarbeiter:
                    tag_mitarbeiter.append(nachname)
                    if is_bulmor and nachname not in tag_bulmor:
                        tag_bulmor.append(nachname)
                elif dienstart == "nacht" and nachname not in nacht_mitarbeiter:
                    nacht_mitarbeiter.append(nachname)
                    if is_bulmor and nachname not in nacht_bulmor:
                        nacht_bulmor.append(nachname)

            # E-Mobby: aus emobby_fahrer.txt + DB-Settings
            try:
                from functions.emobby_functions import is_emobby_fahrer
                tag_emobby   = [m for m in tag_mitarbeiter   if is_emobby_fahrer(m)]
                nacht_emobby = [m for m in nacht_mitarbeiter if is_emobby_fahrer(m)]
            except Exception:
                tag_emobby   = []
                nacht_emobby = []

            self._tag_mitarbeiter   = tag_mitarbeiter
            self._nacht_mitarbeiter = nacht_mitarbeiter
            self._tag_bulmor        = tag_bulmor
            self._nacht_bulmor      = nacht_bulmor
            self._tag_emobby        = tag_emobby
            self._nacht_emobby      = nacht_emobby
            self._dienstplan_geladen = True
            self._btn_open_dienstplan.setEnabled(True)

            # Formular neu aufbauen mit neuen Daten
            self._build_form()

            QMessageBox.information(
                self, "Dienstplan geladen",
                f"Tagdienst: {len(tag_mitarbeiter)} Mitarbeiter "
                f"({len(tag_bulmor)} Bulmor-Fahrer, {len(tag_emobby)} E-Mobby)\n"
                f"Nachtdienst: {len(nacht_mitarbeiter)} Mitarbeiter "
                f"({len(nacht_bulmor)} Bulmor-Fahrer, {len(nacht_emobby)} E-Mobby)"
            )

        except Exception as exc:
            QMessageBox.critical(self, "Fehler", f"Fehler beim Laden:\n{exc}")

    # ── Speichern ───────────────────────────────────────────────────────────

    def _save(self, silent: bool = False) -> Path | None:
        """Sonderaufgaben in Excel speichern."""
        try:
            import openpyxl
        except ImportError:
            if not silent:
                QMessageBox.critical(self, "Fehler", "openpyxl nicht installiert!")
            return None

        if not TEMPLATE_PATH.exists():
            if not silent:
                QMessageBox.critical(
                    self, "Fehler",
                    f"Vorlage nicht gefunden:\n{TEMPLATE_PATH}\n\n"
                    "Bitte Sonderaufgaben.xlsx in Daten/Sonderaufgaben/ ablegen."
                )
            return None

        try:
            datum_heute = datetime.now()
            output_name = f"Sonderaufgaben_{datum_heute.strftime('%Y_%m_%d')}.xlsx"
            output_path = TEMPLATE_PATH.parent / output_name

            copy2(str(TEMPLATE_PATH), str(output_path))

            wb = openpyxl.load_workbook(str(output_path))
            ws = wb.active

            # Datum eintragen
            try:
                ws.cell(row=2, column=1, value=datum_heute.strftime("%d.%m.%Y"))
            except Exception:
                pass

            # Aufgaben-Zeilen schreiben
            for aufgabe, excel_row in _AUFGABEN_MAPPING.items():
                for schicht, col in (("tag", 3), ("nacht", 5)):
                    key = f"{aufgabe}_{schicht}"
                    entry = self._entries.get(key)
                    if entry:
                        val = entry["line"].text().strip()
                        if val:
                            ws.cell(row=excel_row, column=col, value=val)

            # Service Point C72 schreiben
            for key, (row, col) in _SERVICE_MAPPING.items():
                entry = self._entries.get(key)
                if entry:
                    val = entry["line"].text().strip()
                    if val:
                        ws.cell(row=row, column=col, value=val)

            # Bemerkung schreiben
            bemerkung = self._bemerkung.toPlainText().strip()
            if bemerkung:
                ws.cell(row=11, column=1, value=bemerkung)

            wb.save(str(output_path))
            self.reload_tree()   # Baum nach Speichern aktualisieren

            if not silent:
                antwort = QMessageBox.question(
                    self, "Gespeichert",
                    f"Sonderaufgaben gespeichert!\n\nDatei: {output_name}\n\nMöchten Sie die Datei öffnen?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if antwort == QMessageBox.StandardButton.Yes:
                    import os as _os
                    _os.startfile(str(output_path))

            return output_path

        except Exception as exc:
            if not silent:
                QMessageBox.critical(self, "Fehler", f"Fehler beim Speichern:\n{exc}")
            return None

    # ── Excel öffnen ────────────────────────────────────────────────────────

    def _open_excel(self):
        """Öffnet die Excel-Vorlage."""
        if TEMPLATE_PATH.exists():
            try:
                import os as _os
                _os.startfile(str(TEMPLATE_PATH))
            except Exception as exc:
                QMessageBox.critical(self, "Fehler", f"Fehler beim Öffnen:\n{exc}")
        else:
            QMessageBox.critical(
                self, "Fehler",
                f"Vorlage nicht gefunden:\n{TEMPLATE_PATH}"
            )

    # ── Drucken ─────────────────────────────────────────────────────────────

    def _print(self):
        """Zuerst speichern, dann drucken."""
        saved = self._save(silent=True)
        if not saved:
            QMessageBox.warning(
                self, "Fehler",
                "Konnte nicht speichern – bitte erst manuell speichern."
            )
            return

        antwort = QMessageBox.question(
            self, "Drucken",
            f"Datei drucken?\n\n{saved.name}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if antwort == QMessageBox.StandardButton.Yes:
            try:
                import os as _os
                _os.startfile(str(saved), "print")
            except Exception as exc:
                QMessageBox.critical(self, "Fehler", f"Drucken fehlgeschlagen:\n{exc}")

    # ── Dateibaum ────────────────────────────────────────────────────────────

    def _setup_tree(self):
        """Dateibaum für den konfigurierten Sonderaufgaben-Ordner aufbauen."""
        folder = get_setting('sonderaufgaben_ordner', str(TEMPLATE_PATH.parent))
        if not os.path.isdir(folder):
            self._tree_info_lbl.setText("Ordner nicht gefunden.")
            self._tree_info_lbl.setStyleSheet("color: #bb6600; font-size: 10px; padding: 4px;")
            return

        self._fs_model = QFileSystemModel(self)
        self._fs_model.setNameFilters(["*.xlsx", "*.xls"])
        self._fs_model.setNameFilterDisables(False)
        root_idx = self._fs_model.setRootPath(folder)

        self._tree.setModel(self._fs_model)
        self._tree.setRootIndex(root_idx)

        # Nur Dateiname-Spalte anzeigen
        for col in range(1, 4):
            self._tree.hideColumn(col)
        self._tree.header().setVisible(False)

        self._tree_info_lbl.setText(folder)
        self._tree_info_lbl.setStyleSheet("color: #aaa; font-size: 9px; padding: 2px;")

    def reload_tree(self):
        """Baum neu aufbauen (nach Speichern oder manuell)."""
        if self._fs_model is not None:
            self._tree.setModel(None)
            self._fs_model.deleteLater()
            self._fs_model = None
        self._tree_info_lbl.setText("")
        self._setup_tree()

    def _on_tree_activated(self, index):
        """Datei per Doppelklick / Enter als Dienstplan verwenden."""
        if self._fs_model is None:
            return
        path = self._fs_model.filePath(index)
        if os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls")):
            self._parse_dienstplan(path)

    def _tree_context_menu(self, pos):
        """Rechtsklick-Kontextmenü am Dateibaum."""
        if self._fs_model is None:
            return
        index = self._tree.indexAt(pos)
        if not index.isValid():
            return
        path = self._fs_model.filePath(index)
        if not (os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls"))):
            return

        menu = QMenu(self)
        act_dienstplan = menu.addAction("👥 Als Dienstplan verwenden")
        menu.addSeparator()
        act_open = menu.addAction("📄 In Excel öffnen")

        action = menu.exec(self._tree.viewport().mapToGlobal(pos))
        if action == act_dienstplan:
            self._parse_dienstplan(path)
        elif action == act_open:
            try:
                os.startfile(path)
            except Exception as exc:
                QMessageBox.critical(self, "Fehler", f"Fehler beim Öffnen:\n{exc}")

    # ── Neues Formular ────────────────────────────────────────────────────────

    def _new_form(self):
        """Formular zurücksetzen und Mitarbeiter-Listen leeren."""
        self._tag_mitarbeiter   = []
        self._nacht_mitarbeiter = []
        self._tag_bulmor        = []
        self._nacht_bulmor      = []
        self._tag_emobby        = []
        self._nacht_emobby      = []
        self._build_form()

    # ── Refresh (wird beim Tab-Wechsel aufgerufen) ──────────────────────────

    def refresh(self):
        """Baum aktualisieren wenn Tab gewechselt wird."""
        self.reload_tree()
